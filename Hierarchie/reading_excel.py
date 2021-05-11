from loguru import logger
from openpyxl import load_workbook


result = dict()


def check_if_cell_is_id(sheet, column: int, row: int):
    """
    Checks if cell at column and row is an employee id
    :param sheet: excel sheet
    :param column:
    :param row:
    :return:
    """
    value = sheet.cell(column=column, row=row).value
    if value is None:
        return 0
    if value.isdigit():
        return 1
    return 0


def get_stage(column: int, row: int, sheet, decision_maker: str):
    """
    get stage (N-3 for example) and returns it so that the decision maker is filled with the subordinates
    :param column:
    :param row:
    :param sheet: excel sheet
    :param decision_maker:
    :return:
    """
    while column > 1 and check_if_cell_is_id(sheet, column, row) != 1:
        column -= 1
    stage = list()
    while check_if_cell_is_id(sheet, column + 1, row + 1) == 1 or check_if_cell_is_id(sheet, column, row + 1) == 1:
        stage.append(sheet.cell(column=column, row=row).value)
        if check_if_cell_is_id(sheet, column + 1, row + 1) == 1:
            row = get_stage(column=column + 1, row=row + 1, sheet=sheet,
                            decision_maker=sheet.cell(column=column, row=row).value)
            if column > 1 and check_if_cell_is_id(sheet, column, row) != 1:
                result[decision_maker] = stage
                return row
        else:
            row += 1
    if len(stage) == 0 or check_if_cell_is_id(sheet, column, row) == 1:
        stage.append(sheet.cell(column=column, row=row).value)
    result[decision_maker] = stage
    return row + 1


def get_hierarchy_from_excel(path: str):
    excel = load_workbook(path)
    sheet = excel.active
    get_stage(column=1, row=1, sheet=sheet, decision_maker=None)
    for r in result:
        logger.debug(r)
        logger.debug(result[r])
