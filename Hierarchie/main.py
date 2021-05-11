from loguru import logger
import pyodbc
from grhtools import Config
from openpyxl import Workbook

import os
import sys
from PySide6.QtCore import QFile
from PySide6.QtWidgets import QApplication, QMainWindow, QFileDialog, QDialog

from popup import Ui_Dialog
from info_popup import Ui_info
from mainWindow import Ui_MainWindow
from reading_excel import get_hierarchy_from_excel
from call_to_database import get_employees, get_groups, get_society, get_enterprise_from_group

__project_name__ = 'GDS_Hiérarchie'


def fill_decision_maker(employees: dict) -> dict:
    """
    Creates a dict (dec_makers) with decision makers as keys and fills each with their subordinates
    :param employees: all the employees of the group of companies
    :return:
    """
    dec_makers = dict()
    for employee in employees:
        if employees[employee]["idDecideur"] not in dec_makers.keys():
            dec_makers[employees[employee]["idDecideur"]] = list()
        dec_makers[employees[employee]["idDecideur"]].append(employee)
    return dec_makers


def draw_stage(stage: int, column: int, row: int, sheet, employees: dict, decision_makers: dict) -> int:
    """
    Recursive that fills the excel with a complete stage (N-3 for exemple)
    :param stage: id of the employee to write in the file
    :param column: column in which to write
    :param row: row in which to write
    :param sheet: excel sheet in which to write
    :param employees: dict containing all the employees and their information
    :param decision_makers: dict containing every decision_makers in the companies with their subordinate
    :return: the new value of row
    """
    if stage in decision_makers.keys():
        sheet.cell(row=row, column=column, value=str(stage))
        sheet.cell(row=row, column=column + 1, value=employees[stage]['Nom'])
        sheet.cell(row=row, column=column + 2, value=employees[stage]['Prenom'])
        sheet.cell(row=row, column=column + 3, value=employees[stage]['CodMatricule'])
        for s in decision_makers[stage]:
            row = draw_stage(s, column + 1, row + 1, sheet, employees, decision_makers)
    else:
        sheet.cell(row=row, column=column, value=str(stage))
        sheet.cell(row=row, column=column + 1, value=employees[stage]['Nom'])
        sheet.cell(row=row, column=column + 2, value=employees[stage]['Prenom'])
        sheet.cell(row=row, column=column + 3, value=employees[stage]['CodMatricule'])
        if column == 1:
            row += 1
    return row


def draw_hierarchy(decision_makers: dict, employees: dict, path: str):
    """
    creates and fills an excel file with the hierarchy of the group of companies or company.
    :param decision_makers: dict with all the decision_makers of the companies as keys and their subordinates as values
    :param employees: all the employees in the companies and their information
    :param path: path for the excel saving
    :return:
    """
    excel = Workbook()
    sheet = excel.active
    column = 1
    row = 1
    stage = decision_makers[None]
    for s in stage:
        if employees[s]['idEmploi'] != 28:
            row = draw_stage(s, column, row, sheet, employees=employees, decision_makers=decision_makers)
        else:
            row += 1
            sheet.cell(row=row, column=column, value=str(s))
            sheet.cell(row=row, column=column + 1, value=employees[s]['Nom'])
            sheet.cell(row=row, column=column + 2, value=employees[s]['Prenom'])
            sheet.cell(row=row, column=column + 3, value=employees[s]['CodMatricule'])
    excel.save(path)


class InformationPopup(QDialog):
    def __init__(self, parent=None):
        super(InformationPopup, self).__init__(parent)
        self.ui = Ui_info()
        self.ui.setupUi(self)
        self.setWindowTitle("Information popup")

        self.ui.pushButton.clicked.connect(self.ok_button)

    def set_information(self, information):
        self.ui.textBrowser.setText(information)

    def ok_button(self):
        self.close()


class PopUp(QDialog):
    def __init__(self, parent=None):
        super(PopUp, self).__init__(parent)
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        self.setWindowTitle("Popup")
        self.status = -1

        self.ui.cancel_button.clicked.connect(self.cancel_btn_clicked)
        self.ui.validation_button.clicked.connect(self.validation_btn_clicked)

    def set_question(self, question):
        self.ui.question_textBrowser.setText(question)

    def validation_btn_clicked(self):
        self.status = 1
        self.close()

    def cancel_btn_clicked(self):
        self.status = -1
        self.close()


class MainWindow(QMainWindow):
    def __init__(self, cursor, defautl_path):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setWindowTitle("Hiérarchie")
        self.popup = PopUp()
        self.information = InformationPopup()

        self.cursor = cursor
        self.excel_path = ""
        self.default_path = defautl_path

        self.groups = get_groups(cursor_pwa, config.config['black_list']['group'])
        self.setup_groups()
        self.companies = get_society(cursor_pwa, config.config['black_list']['enterprise'])
        self.companies = dict(sorted(self.companies.items(), key=lambda x: x[1]['NomSociete']))
        self.setup_companies()

        self.ui.group_comboBox.currentIndexChanged.connect(self.group_dropwdown_changed)
        self.ui.getHierarchy_button.clicked.connect(self.get_hierarchy)
        self.ui.get_excel_file_button.clicked.connect(self.get_excel_file)
        self.ui.set_hierarchy_button.clicked.connect(self.set_hierarchy)

    def setup_groups(self):
        self.ui.group_comboBox.addItem("")
        for g in self.groups:
            self.ui.group_comboBox.addItem(g)

    def setup_companies(self):
        self.ui.companies_comboBox.addItem("")
        for c in self.companies:
            self.ui.companies_comboBox.addItem(self.companies[c]["NomSociete"])

    def group_dropwdown_changed(self, index: int):
        if index > 0:
            self.ui.companies_comboBox.setEnabled(False)
        else:
            self.ui.companies_comboBox.setEnabled(True)

    def get_hierarchy(self):
        """
        function that gathers the hierarchy of a company or a group of companies.
        Asking for the directory of destination
        call the database to get the employees
        creates a dict with the decision makers and their subordinates from the dict self.employees
        then calls draw hierarchy to create the excel file
        :return:
        """
        if self.ui.group_comboBox.currentIndex() == 0 and self.ui.companies_comboBox.currentIndex() == 0:
            self.information.set_information(
                "Veuillez choisir une entreprise ou un groupe")
            self.information.exec()
            return
        directory = QFileDialog.getExistingDirectory(self, "Choix du dossier", self.default_path)
        if len(directory) == 0:
            return
        if self.ui.group_comboBox.currentIndex() > 0:
            employees = get_employees(self.cursor, get_enterprise_from_group(self.cursor, self.groups[self.ui.group_comboBox.currentText()]['idGroupe']))
        else:
            employees = get_employees(self.cursor, self.get_id_company())
        decision_makers = fill_decision_maker(employees=employees)
        if self.ui.group_comboBox.currentIndex() != 0:
            draw_hierarchy(decision_makers=decision_makers, employees=employees,
                           path=self.ui.group_comboBox.currentText() + "-hiérarchie.xlsx")
        else:
            draw_hierarchy(decision_makers=decision_makers, employees=employees,
                           path=self.ui.companies_comboBox.currentText() + "-hiérarchie.xlsx")
        self.ui.group_comboBox.setCurrentIndex(0)
        self.ui.companies_comboBox.setCurrentIndex(0)
        self.information.set_information("La hiérarchie a été récupérée")
        self.information.exec()

    def get_id_company(self):
        for c in self.companies:
            if self.companies[c]['NomSociete'] == self.ui.companies_comboBox.currentText():
                return [c]

    def get_excel_file(self):
        path = QFileDialog.getOpenFileName(self, "Choix du fichier excel", self.default_path, "Fichier excel (*.xlsx)")[
            0]
        logger.debug(path)
        self.excel_path = path
        if path == "":
            self.ui.get_excel_file_button.setText("Choisir le fichier excel")
        else:
            self.ui.get_excel_file_button.setText(os.path.split(path)[1])

    def set_hierarchy(self):
        if self.excel_path != "":
            self.information.set_information(
                "Veuillez choisir le fichier excel contenant la hiérarchie pour pouvoir la changer")
            self.information.exec()
        else:
            self.popup.set_question("Voulez-vous changer la hiérarchie ?")
            self.popup.exec()
            if self.popup.status == 1:
                get_hierarchy_from_excel(self.excel_path)
                self.ui.get_excel_file_button.setText("Choisir le fichier excel")
                self.information.set_information("La hiérarchie a été mise à jour")
            self.information.exec()


if __name__ == '__main__':
    config = Config(project_name=__project_name__)
    cursor_pwa = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + config.config['login_pwa']['server']
                                + ';DATABASE=' + config.config['login_pwa']['database'] + ';UID=' +
                                config.config['login_pwa']['username'] + ';PWD=' +
                                config.config['login_pwa']['password']).cursor()

    app = QApplication(sys.argv)
    window = MainWindow(cursor=cursor_pwa, defautl_path=config.config['path'])
    window.show()

    sys.exit(app.exec_())
