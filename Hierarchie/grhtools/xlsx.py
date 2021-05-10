# -*- coding: utf-8 -*-
from loguru import logger
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import TableStyleInfo, Table


font_colors = {
    'RED': '9C0006',
    'GREEN': '006100',
    'YELLOW': '9C5700',
    'WHITE': 'FFFFFF',
    'BLACK': '000000',
}

fill_colors = {
    'RED': 'FFC7CE',
    'GREEN': 'C6EFCE',
    'YELLOW': 'FFEB9C',
    'WHITE': 'FFFFFF',
    'BLACK': '000000',
}


def get_font(color_name, font='Calibri', bold=False, italic=False, color=None):
    """
    :param color_name: Color name, (RED, GREEN, YELLOW, WHITE)
    :param font: Font name, default: Calibri
    :param bold: Boolean
    :param italic: Boolean
    :param color: RGB without # only if custom color name not in list
    :return: openpyxl.styles.Font
    """
    if color:
        return Font(name=font, color=color, bold=bold, italic=italic)
    elif color_name in font_colors:
        return Font(name=font, color=font_colors[color_name], bold=bold, italic=italic)
    else:
        return Font(name=font, bold=bold, italic=italic)


def get_patternfill(color_name, patern_type="solid", color=None):
    """
    :param color_name: Color name, (RED, GREEN, YELLOW, WHITE)
    :param patern_type:
    :param color: RGB without # only if custom color name not in list
    :return:
    """
    if color:
        return PatternFill(patern_type, fgColor=color)
    elif color_name in fill_colors:
        return PatternFill(patern_type, fgColor=fill_colors[color_name])
    else:
        return PatternFill(patern_type)


def get_borders(color='000000', thick=False):
    if color != '000000':
        if color in fill_colors:
            color = fill_colors[color]
    if thick is False:
        size = Side(border_style='thin', color=color)
    else:
        size = Side(border_style='thick', color=color)
    return Border(left=size, right=size, top=size, bottom=size)


def get_alignment(horizontal='left', vertical='center', wrap_text=False, text_rotation=0):
    return Alignment(horizontal=horizontal, vertical=vertical, textRotation=text_rotation, wrap_text=wrap_text)


def set_success_colors(cell):
    cell.font = get_font(None, color='006100')
    cell.fill = get_patternfill(None, color='C6EFCE')


def set_error_colors(cell):
    cell.font = get_font(None, color='9C0006')
    cell.fill = get_patternfill(None, color='FFC7CE')


def set_warning_colors(cell):
    cell.font = get_font(None, color='9C5700')
    cell.fill = get_patternfill(None, color='FFEB9C')


def set_specific_colors(cell, fill, font="000000"):
    cell.font = get_font(None, color=font)
    cell.fill = get_patternfill(None, color=fill)


def set_alignment(cell, horizontal, vertical='center', wrap_text=False, text_rotation=0):
    cell.alignment = get_alignment(horizontal, vertical, wrap_text=wrap_text, text_rotation=text_rotation)


def set_print_settings(sheet, orientation='portrait', papersize='A4', title_to_print='$1:$1', auto_style=True,
                       header_left='&D &T', header_center='&F', header_right='&P / &N', freeze_panes=None,
                       table_style='TableStyleMedium2', table_name='Table', autostart="A1"):
    """

    :param sheet: targeted Worksheet
    :param orientation: paper orientation
    :param papersize: A3 A4 ...
    :param title_to_print: title rows that wiil b print on each page
    :param auto_style: will create a Medium 2 table with autofilter an odd/even coloring
    :param header_left: text on left header of each page
    :param header_center: text on center header of each page
    :param header_right: text on right header of each page
    :param freeze_panes: cell to freez panes
    :param table_style: MS Excell styles name
    :param table_name: MS Excell table name comment
    """
    if auto_style:
        medium_style = TableStyleInfo(table_style, showRowStripes=True)
        table = Table(ref=f'{autostart}:{get_column_letter(sheet.max_column)}{sheet.max_row}',
                      tableStyleInfo=medium_style, displayName=table_name)
        sheet.add_table(table)

    if freeze_panes:
        sheet.freeze_panes = freeze_panes

    sheet.oddHeader.left.text = header_left
    sheet.oddHeader.center.text = header_center
    sheet.oddHeader.right.text = header_right
    if title_to_print:
        sheet.print_title_rows = title_to_print
    # sheet s.print_area = ['$A:$B']
    sheet.print_options.horizontalCentered = True
    sheet.print_options.verticalCentered = False
    if orientation == 'portrait':
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    elif orientation == 'landscape':
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.page_setup.fitToHeight = False
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.50
    sheet.page_margins.bottom = 0
    sheet.page_margins.footer = 0
    sheet.page_margins.header = 0
    sheet.page_setup.paperSize = getattr(sheet, 'PAPERSIZE_'+papersize.upper())


if __name__ == '__main__':
    from openpyxl import Workbook

    w = Workbook()
    s = w.active

    for col in range(1, 15):
        s.cell(row=1, column=col).value = f'lorem ipsum dolor sit amet {col}'
        set_alignment(s.cell(row=1, column=col), 'center', text_rotation=45)
        for row in range(2, 100):
            s.cell(row=row, column=col).value = 'lorem ipsum dolor sit amet'
            s.cell(row=row, column=col).border = get_borders(color='7e57c2')
            set_alignment(s.cell(row=row, column=col), 'left', wrap_text=True)

    s.cell(row=1, column=2).font = get_font('RED')
    s.cell(row=2, column=2).font = get_font('GREEN')
    s.cell(row=3, column=2).font = get_font('YELLOW', bold=True)
    s.cell(row=4, column=2).font = get_font(None, italic=True, color='004BA0')

    s.cell(row=1, column=2).fill = get_patternfill('RED')
    s.cell(row=2, column=2).fill = get_patternfill('GREEN')
    s.cell(row=3, column=2).fill = get_patternfill('YELLOW')
    s.cell(row=4, column=2).fill = get_patternfill(None, color='1976D2')

    set_success_colors(s.cell(row=6, column=2))
    set_error_colors(s.cell(row=7, column=2))
    set_warning_colors(s.cell(row=8, column=2))

    set_print_settings(s)

    w.save('test.xlsx')
